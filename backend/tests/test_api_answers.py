"""Integration tests for the AI-powered answers API."""

import pytest


@pytest.mark.asyncio
async def test_generate_answers_no_questions(client, auth_headers):
    resp = await client.post(
        "/api/v1/answers/",
        json={"questions": []},
        headers=auth_headers,
    )
    assert resp.status_code == 400
    assert "at least one" in resp.json()["detail"].lower()


@pytest.mark.asyncio
async def test_generate_answers_with_evidence(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Encryption Policy",
            "type": "policy",
            "frameworks": ["soc2"],
            "control_ids": ["CC6.1"],
            "last_reviewed": "2026-01-15",
            "owner": "security@test.com",
            "summary": "We encrypt all data at rest and in transit.",
            "snippets": [
                "All customer data is encrypted at rest using AES-256.",
                "Data in transit uses TLS 1.3 exclusively.",
            ],
        },
        headers=auth_headers,
    )

    resp = await client.post(
        "/api/v1/answers/",
        json={
            "questions": ["How do you encrypt customer data?"],
            "use_llm": False,
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201
    data = resp.json()
    assert len(data["answers"]) == 1
    answer = data["answers"][0]
    assert answer["question"] == "How do you encrypt customer data?"
    assert len(answer["answer_text"]) > 0
    assert answer["confidence"] in ("high", "medium", "low")
    assert len(answer["citations"]) >= 0


@pytest.mark.asyncio
async def test_generate_answers_multiple_questions(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Access Control Policy",
            "type": "policy",
            "frameworks": ["soc2"],
            "control_ids": ["CC6.2"],
            "last_reviewed": "2026-01-15",
            "owner": "security@test.com",
            "summary": "MFA enforced for all systems.",
            "snippets": ["All production systems require MFA."],
        },
        headers=auth_headers,
    )

    resp = await client.post(
        "/api/v1/answers/",
        json={
            "questions": [
                "How do you handle access control?",
                "Do you support SAML SSO?",
            ],
            "use_llm": False,
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201
    data = resp.json()
    assert len(data["answers"]) == 2
    assert data["engine_used"] in ("ai", "bm25")


@pytest.mark.asyncio
async def test_generate_answers_no_evidence(client, auth_headers):
    resp = await client.post(
        "/api/v1/answers/",
        json={
            "questions": ["How do you encrypt data?"],
            "use_llm": False,
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201
    data = resp.json()
    assert len(data["answers"]) == 1
    assert data["confidence_counts"]["low"] == 1


@pytest.mark.asyncio
async def test_list_answer_generations(client, auth_headers):
    resp = await client.get("/api/v1/answers/", headers=auth_headers)
    assert resp.status_code == 200
    assert isinstance(resp.json(), list)


@pytest.mark.asyncio
async def test_get_answer_generation_not_found(client, auth_headers):
    resp = await client.get("/api/v1/answers/nonexistent-id", headers=auth_headers)
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_assign_answer(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    gen_data = gen_resp.json()
    answer_id = gen_data["answers"][0]["id"]

    me_resp = await client.get("/api/v1/auth/me", headers=auth_headers)
    user_id = me_resp.json()["id"]

    resp = await client.post(
        "/api/v1/answers/assign",
        json={"answer_id": answer_id, "assignee_id": user_id},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["assignee_id"] == user_id


@pytest.mark.asyncio
async def test_bulk_assign_answers(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Q1?", "Q2?"], "use_llm": False},
        headers=auth_headers,
    )
    gen_data = gen_resp.json()
    answer_ids = [a["id"] for a in gen_data["answers"]]

    me_resp = await client.get("/api/v1/auth/me", headers=auth_headers)
    user_id = me_resp.json()["id"]

    resp = await client.post(
        "/api/v1/answers/bulk-assign",
        json={"answer_ids": answer_ids, "assignee_id": user_id},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["assigned_count"] == 2


@pytest.mark.asyncio
async def test_list_assigned_answers(client, auth_headers):
    resp = await client.get("/api/v1/answers/assigned", headers=auth_headers)
    assert resp.status_code == 200
    assert isinstance(resp.json(), list)


@pytest.mark.asyncio
async def test_update_answer(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    answer_id = gen_resp.json()["answers"][0]["id"]

    resp = await client.put(
        f"/api/v1/answers/{answer_id}",
        json={"answer_text": "Updated answer text", "needs_human_review": False},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["answer_text"] == "Updated answer text"
    assert resp.json()["needs_human_review"] is False


@pytest.mark.asyncio
async def test_regenerate_single_answer(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    answer_id = gen_resp.json()["answers"][0]["id"]

    resp = await client.post(
        f"/api/v1/answers/regenerate/{answer_id}",
        json={"questions": ["New question?"], "use_llm": False},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["question"] == "New question?"


@pytest.mark.asyncio
async def test_learn_from_approvals(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    answer_ids = [a["id"] for a in gen_resp.json()["answers"]]

    for aid in answer_ids:
        await client.put(
            f"/api/v1/answers/{aid}",
            json={"answer_text": "Approved answer text", "needs_human_review": False},
            headers=auth_headers,
        )
        await client.post(
            "/api/v1/approvals/",
            json={"question": "Test question?", "status": "approved", "notes": "Looks good"},
            headers=auth_headers,
        )

    resp = await client.post(
        "/api/v1/answers/learn",
        json={"answer_ids": answer_ids},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["learned"] >= 0


@pytest.mark.asyncio
async def test_knowledge_base_search(client, auth_headers):
    resp = await client.get(
        "/api/v1/answers/knowledge-base/search",
        params={"q": "encryption"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert "results" in resp.json()


@pytest.mark.asyncio
async def test_questionnaire_crud(client, auth_headers):
    resp = await client.post(
        "/api/v1/answers/questionnaires",
        json={
            "name": "Acme Corp Security Review",
            "questions": ["How do you encrypt data?", "Do you have MFA?"],
            "original_filename": "acme-review.xlsx",
            "original_format": "xlsx",
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201
    q_data = resp.json()
    assert q_data["name"] == "Acme Corp Security Review"
    assert q_data["question_count"] == 2
    questionnaire_id = q_data["id"]

    list_resp = await client.get("/api/v1/answers/questionnaires", headers=auth_headers)
    assert list_resp.status_code == 200
    assert len(list_resp.json()) >= 1

    update_resp = await client.put(
        f"/api/v1/answers/questionnaires/{questionnaire_id}",
        json={"status": "completed"},
        headers=auth_headers,
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["status"] == "completed"


@pytest.mark.asyncio
async def test_import_file(client, auth_headers):
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="Question")
    ws.cell(row=1, column=2, value="Response")
    ws.cell(row=2, column=1, value="Do you encrypt data at rest?")
    ws.cell(row=3, column=1, value="Is MFA enforced for all users?")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    files = {"file": ("test.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

    resp = await client.post(
        "/api/v1/answers/import-file",
        files=files,
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["count"] >= 1
    assert data["format"] == "xlsx"


@pytest.mark.asyncio
async def test_export_questionnaire_xlsx(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    gen_id = gen_resp.json()["id"]

    resp = await client.post(
        "/api/v1/export/questionnaire/xlsx",
        json={"generation_id": gen_id, "format": "xlsx"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_export_questionnaire_docx(client, auth_headers):
    await client.post(
        "/api/v1/evidence/",
        json={
            "title": "Test Evidence",
            "type": "test",
            "frameworks": [],
            "control_ids": [],
            "last_reviewed": "2026-01-15",
            "owner": "tester@test.com",
            "summary": "Test evidence.",
            "snippets": ["Test snippet."],
        },
        headers=auth_headers,
    )

    gen_resp = await client.post(
        "/api/v1/answers/",
        json={"questions": ["Test question?"], "use_llm": False},
        headers=auth_headers,
    )
    gen_id = gen_resp.json()["id"]

    resp = await client.post(
        "/api/v1/export/questionnaire/docx",
        json={"generation_id": gen_id, "format": "docx"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert "wordprocessingml" in resp.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_sample_questions(client, auth_headers):
    resp = await client.get("/api/v1/answers/sample", headers=auth_headers)
    assert resp.status_code == 200
    assert "questions" in resp.json()
